from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from datetime import datetime
import io
import re
import openpyxl
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from xml.sax.saxutils import escape as xml_escape
from .. import models, schemas, database, auth
from .logs import log_activity


def _attachment_header(stem: str, ext: str) -> dict:
    """
    A Content-Disposition a browser will accept.

    `filename=schedule_College_of_Arts,_Sciences_&_Technology.pdf` was being sent
    unquoted, and a comma is a header-value separator: Chrome read it as two
    Content-Disposition headers and refused the whole response
    (ERR_RESPONSE_HEADERS_MULTIPLE_CONTENT_DISPOSITION). Every college at DLSAU
    has a comma in its name, so the PDF and Excel exports failed for all of them.
    Reduced to a safe ASCII stem, and quoted.
    """
    safe = re.sub(r'[^A-Za-z0-9._-]+', '_', stem).strip('_') or 'schedule'
    return {"Content-Disposition": f'attachment; filename="{safe}.{ext}"'}

router = APIRouter(
    prefix="/api/schedules",
    tags=["Schedules"]
)


def _resolve_scope_department(db: Session, current_user: models.User, requested_id: Optional[int]):
    """
    The department a request should be scoped to.

    Chairs and coordinators are pinned to their own department regardless of
    what they ask for. Admins may target one explicitly, or None for every
    department. Returns None only for an unscoped admin.
    """
    if current_user.role in ('program_chair', 'coordinator'):
        dept = db.query(models.Department).filter(
            (models.Department.code == current_user.department) |
            (models.Department.name == current_user.department)
        ).first()
        if not dept:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Your department could not be resolved."
            )
        return dept.id
    return requested_id

@router.get("", response_model=List[schemas.ScheduleResponse])
def get_schedules(
    skip: int = 0, 
    limit: int = 100, 
    semester_id: Optional[int] = None,
    curriculum_id: Optional[int] = None,
    faculty_id: Optional[int] = None,
    room_id: Optional[int] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    query = db.query(models.Schedule).options(
        joinedload(models.Schedule.curriculum),
        joinedload(models.Schedule.room)
    ).join(models.Curriculum)
    
    if current_user.role in ['program_chair', 'coordinator']:
        if not current_user.department:
            return []
        dept = db.query(models.Department).filter(
            (models.Department.code == current_user.department) | 
            (models.Department.name == current_user.department)
        ).first()
        if dept:
            query = query.filter(models.Curriculum.department_id == dept.id)
        else:
            return []
            
    if semester_id:
        query = query.filter(models.Schedule.semester_id == semester_id)
    if curriculum_id:
        query = query.filter(models.Schedule.curriculum_id == curriculum_id)
    if faculty_id:
        query = query.filter(models.Schedule.faculty_id == faculty_id)
    if room_id:
        query = query.filter(models.Schedule.room_id == room_id)
        
    return query.offset(skip).limit(limit).all()

# NOTE: GET /{schedule_id} is deliberately registered near the bottom of this
# module. FastAPI matches routes in registration order, so declaring it here
# would swallow every literal sibling path -- /suggestions used to return a 422
# "unable to parse 'suggestions' as an integer" for exactly that reason.
# Keep literal paths above the catch-all parameter route.

@router.post("", response_model=schemas.ScheduleResponse, status_code=status.HTTP_201_CREATED)
def create_schedule(
    schedule: schemas.ScheduleCreate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role not in ['program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Administrators do not create classes. This is done by the program chair or coordinator.")
        
    curriculum_item = db.query(models.Curriculum).filter(models.Curriculum.id == schedule.curriculum_id).first()
    if not curriculum_item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Curriculum item not found")

    if current_user.role in ['program_chair', 'coordinator']:
        dept = db.query(models.Department).filter(models.Department.id == curriculum_item.department_id).first()
        if not dept or (dept.code != current_user.department and dept.name != current_user.department):
             raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Can only create schedules for your department's curriculum")
             
    new_schedule = models.Schedule(**schedule.model_dump())
    db.add(new_schedule)
    db.commit()
    db.refresh(new_schedule)
    dept_id_val = getattr(curriculum_item, 'department_id', None)
    log_activity(db, current_user.id, "Create Schedule", f"Created schedule for {curriculum_item.code} in {new_schedule.section}", "success", department_id=dept_id_val) # type: ignore
    
    return new_schedule

@router.put("/{schedule_id}", response_model=schemas.ScheduleResponse)
def update_schedule(
    schedule_id: int, 
    schedule: schemas.ScheduleUpdate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role not in ['program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Administrators do not edit classes. This is done by the program chair or coordinator.")
        
    db_schedule = db.query(models.Schedule).filter(models.Schedule.id == schedule_id).first()
    if not db_schedule:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schedule not found")
        
    if current_user.role in ['program_chair', 'coordinator']:
        curriculum_item = db.query(models.Curriculum).filter(models.Curriculum.id == db_schedule.curriculum_id).first()
        if curriculum_item:
            dept = db.query(models.Department).filter(models.Department.id == curriculum_item.department_id).first()
            if not dept or (dept.code != current_user.department and dept.name != current_user.department):
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized to modify this schedule")
    
    if db_schedule.is_locked and schedule.is_locked is not False:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Schedule is locked and cannot be modified")
                
    update_data = schedule.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_schedule, key, value)
        
    db.commit()
    db.refresh(db_schedule)
    
    curriculum_item = db.query(models.Curriculum).filter(models.Curriculum.id == db_schedule.curriculum_id).first()
    item_code = curriculum_item.code if curriculum_item else "Unknown"
    log_activity(db, current_user.id, "Update Schedule", f"Updated schedule for {item_code} in {db_schedule.section}", "success", department_id=curriculum_item.department_id if curriculum_item else None) # type: ignore
    
    return db_schedule

@router.delete("/clear-all")
def clear_all_schedules(
    semester_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role not in ['program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Administrators do not clear schedules. This is done by the program chair or coordinator.")

    query = db.query(models.Schedule).filter(models.Schedule.semester_id == semester_id)

    if current_user.role in ['program_chair', 'coordinator']:
        if not current_user.department:
            return {"deleted_count": 0, "backup": []}
        dept = db.query(models.Department).filter(
            (models.Department.code == current_user.department) |
            (models.Department.name == current_user.department)
        ).first()
        if dept:
            query = query.join(models.Curriculum).filter(models.Curriculum.department_id == dept.id)
        else:
            return {"deleted_count": 0, "backup": []}

    schedules_to_delete = query.all()
    backup = []
    for s in schedules_to_delete:
        backup.append({
            "semester_id": s.semester_id,
            "curriculum_id": s.curriculum_id,
            "faculty_id": s.faculty_id,
            "room_id": s.room_id,
            "day_of_week": s.day_of_week,
            "start_time": s.start_time.strftime("%H:%M") if s.start_time else "07:30",
            "end_time": s.end_time.strftime("%H:%M") if s.end_time else "09:00",
            "section": s.section or "",
            "status": s.status or "draft"
        })

    count = len(schedules_to_delete)
    for s in schedules_to_delete:
        db.delete(s)
    db.commit()

    log_activity(db, current_user.id, "Clear All Schedules", f"Cleared {count} schedules for semester ID {semester_id}", "success")
    return {"deleted_count": count, "backup": backup}

from pydantic import BaseModel

class ScheduleStatusRequest(BaseModel):
    semester_id: int
    department_id: Optional[int] = None
    status: str


@router.get("/status")
def get_schedule_status(
    semester_id: int,
    department_id: Optional[int] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Publication state for one department's schedule in one term.

    A schedule is Published only when every class in it is published; a single
    draft class means the term is still a draft. Anything else would let a
    department post a schedule on its door while part of it was still unsaved.
    """
    if current_user.role not in ['admin', 'program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")

    dept_id = _resolve_scope_department(db, current_user, department_id)

    query = db.query(models.Schedule).join(models.Curriculum).filter(
        models.Schedule.semester_id == semester_id
    )
    if dept_id is not None:
        query = query.filter(models.Curriculum.department_id == dept_id)

    rows = query.all()
    total = len(rows)
    published = sum(1 for s in rows if s.status == 'published')

    unresolved = db.query(models.Conflict).filter(models.Conflict.resolved_at == None).count()

    return {
        "semester_id": semester_id,
        "department_id": dept_id,
        "total": total,
        "published": published,
        "status": "published" if total > 0 and published == total else "draft",
        "unresolved_conflicts": unresolved,
    }


@router.patch("/status")
def set_schedule_status(
    payload: ScheduleStatusRequest,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Publish or unpublish a department's schedule for a term (DEP-1).

    Publishing is what makes a schedule official, so it is an administrator
    action and it is refused while conflicts are unresolved -- posting a
    timetable that the system knows is broken is the failure this prevents.
    """
    if current_user.role != 'admin':
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only administrators can publish schedules."
        )

    new_status = (payload.status or '').strip().lower()
    if new_status not in ('draft', 'published'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Status must be either draft or published."
        )

    semester = db.query(models.Semester).filter(models.Semester.id == payload.semester_id).first()
    if not semester:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Term not found")

    dept_id = payload.department_id
    dept = db.query(models.Department).filter(models.Department.id == dept_id).first() if dept_id else None

    query = db.query(models.Schedule).join(models.Curriculum).filter(
        models.Schedule.semester_id == payload.semester_id
    )
    if dept_id is not None:
        query = query.filter(models.Curriculum.department_id == dept_id)

    rows = query.all()
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="There is nothing to publish for this term. Generate a schedule first."
        )

    if new_status == 'published':
        unresolved = db.query(models.Conflict).filter(models.Conflict.resolved_at == None).count()
        if unresolved:
            noun = "conflict" if unresolved == 1 else "conflicts"
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    f"{unresolved} unresolved {noun} must be cleared before this schedule "
                    "can be published."
                )
            )

    for s in rows:
        s.status = new_status
    db.commit()

    dept_name = dept.name if dept else "All departments"
    verb = "Publish Schedule" if new_status == 'published' else "Unpublish Schedule"
    log_activity(
        db,
        current_user.id,  # type: ignore
        verb,
        f"{dept_name} · {semester.academic_year} {semester.term} · {len(rows)} classes set to {new_status}",
        "success",
        department_id=dept_id
    )

    return {"status": new_status, "affected": len(rows), "department_id": dept_id}


class RestoreSchedulesRequest(BaseModel):
    items: List[schemas.ScheduleCreate]

@router.post("/restore", status_code=status.HTTP_201_CREATED)
def restore_schedules(
    req: RestoreSchedulesRequest,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role not in ['program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Administrators do not restore schedules. This is done by the program chair or coordinator.")

    # Restore is a bulk insert, so it needs the same department check the
    # single-item create path applies -- otherwise it is a way to write
    # schedules into any department's curriculum.
    allowed_dept_id = None
    if current_user.role in ['program_chair', 'coordinator']:
        dept = db.query(models.Department).filter(
            (models.Department.code == current_user.department) |
            (models.Department.name == current_user.department)
        ).first()
        if not dept:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Your department could not be resolved")
        allowed_dept_id = dept.id

    restored_items = []
    for item in req.items:
        if allowed_dept_id is not None:
            curriculum_item = db.query(models.Curriculum).filter(
                models.Curriculum.id == item.curriculum_id
            ).first()
            if not curriculum_item or curriculum_item.department_id != allowed_dept_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail=f"Cannot restore a schedule for curriculum {item.curriculum_id}: it belongs to another department"
                )
        new_sched = models.Schedule(**item.model_dump())
        db.add(new_sched)
        restored_items.append(new_sched)

    db.commit()
    log_activity(db, current_user.id, "Restore Schedules", f"Restored {len(restored_items)} schedule items", "success")
    return {"restored_count": len(restored_items)}

@router.delete("/{schedule_id}")
def delete_schedule(
    schedule_id: int, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role not in ['program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Administrators do not delete classes. This is done by the program chair or coordinator.")
        
    db_schedule = db.query(models.Schedule).filter(models.Schedule.id == schedule_id).first()
    if not db_schedule:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schedule not found")
        
    if current_user.role in ['program_chair', 'coordinator']:
        curriculum_item = db.query(models.Curriculum).filter(models.Curriculum.id == db_schedule.curriculum_id).first()
        if curriculum_item:
            dept = db.query(models.Department).filter(models.Department.id == curriculum_item.department_id).first()
            if not dept or (dept.code != current_user.department and dept.name != current_user.department):
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized to delete this schedule")
    
    if db_schedule.is_locked:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Locked schedules cannot be deleted")
                
    curriculum_item = db.query(models.Curriculum).filter(models.Curriculum.id == db_schedule.curriculum_id).first()
    section = db_schedule.section
    dept_id = curriculum_item.department_id if curriculum_item else None
    code = curriculum_item.code if curriculum_item else "Unknown"

    backup = {
        "semester_id": db_schedule.semester_id,
        "curriculum_id": db_schedule.curriculum_id,
        "faculty_id": db_schedule.faculty_id,
        "room_id": db_schedule.room_id,
        "day_of_week": db_schedule.day_of_week,
        "start_time": db_schedule.start_time.strftime("%H:%M") if db_schedule.start_time else "07:30",
        "end_time": db_schedule.end_time.strftime("%H:%M") if db_schedule.end_time else "09:00",
        "section": db_schedule.section or "",
        "status": db_schedule.status or "draft"
    }

    db.delete(db_schedule)
    db.commit()
    
    log_activity(db, current_user.id, "Delete Schedule", f"Deleted schedule for {code} in {section}", "success", department_id=dept_id) # type: ignore
    
    return {"message": "Schedule deleted successfully", "backup": backup}

@router.get("/suggestions")
def get_schedule_suggestions(
    curriculum_id: int,
    semester_id: Optional[int] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Returns a list of conflict-free scheduling suggestions for a given curriculum item.
    """
    if current_user.role not in ['admin', 'program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
        
    curriculum_item = db.query(models.Curriculum).filter(models.Curriculum.id == curriculum_id).first()
    if not curriculum_item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Curriculum item not found")
        
    if current_user.role in ['program_chair', 'coordinator']:
        dept = db.query(models.Department).filter(models.Department.id == curriculum_item.department_id).first()
        if not dept or (dept.code != current_user.department and dept.name != current_user.department):
             raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized for this curriculum item")

    # If semester_id not provided, try to find active semester
    if not semester_id:
        active_sem = db.query(models.Semester).filter(models.Semester.is_active == True).first()
        if not active_sem:
            return []
        semester_id = active_sem.id # type: ignore

    from datetime import time
    TIMESLOTS = [
        (time(8, 0), time(9, 30)),
        (time(9, 30), time(11, 0)),
        (time(11, 0), time(12, 30)),
        (time(13, 0), time(14, 30)),
        (time(14, 30), time(16, 0)),
        (time(16, 0), time(17, 30)),
    ]
    DAYS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

    # Get valid rooms for the curriculum item type
    rooms = db.query(models.Room).all()
    valid_rooms = [r for r in rooms if r.type == curriculum_item.type or (r.type == 'computer_lab' and curriculum_item.type == 'lab')]

    # Get valid faculties for the curriculum department
    faculties = db.query(models.Faculty).filter(models.Faculty.department_id == curriculum_item.department_id).all()
    valid_faculties = []
    
    # One pass for the whole semester instead of a query per faculty plus a
    # query per schedule row. `semester_schedules` is reused below for the
    # overlap checks, which used to issue their own query per candidate slot.
    semester_schedules = db.query(models.Schedule).filter(
        models.Schedule.semester_id == semester_id
    ).all()
    unit_by_curriculum = {
        c.id: int(c.units or 0)
        for c in db.query(models.Curriculum.id, models.Curriculum.units).all()
    }
    units_by_faculty = {}
    for s in semester_schedules:
        if s.faculty_id:
            units_by_faculty[s.faculty_id] = (
                units_by_faculty.get(s.faculty_id, 0) + unit_by_curriculum.get(s.curriculum_id, 0)
            )

    # Faculty unavailability, likewise loaded once.
    unavailabilities = db.query(models.FacultyUnavailability).filter(
        models.FacultyUnavailability.faculty_id.in_([f.id for f in faculties])
    ).all() if faculties else []

    # Calculate current units for faculties
    for f in faculties:
        total_units = units_by_faculty.get(f.id, 0)
        if (total_units + curriculum_item.units) <= f.max_units:
            valid_faculties.append({
                "faculty": f,
                "load_percentage": (total_units + curriculum_item.units) / f.max_units
            })

    suggestions = []
    
    # Limit suggestions to top 5 to keep it fast and relevant
    import random
    random.shuffle(valid_rooms)
    
    # Sort faculties by load percentage (lowest load first)
    valid_faculties.sort(key=lambda x: x["load_percentage"])

    for f_data in valid_faculties[:3]:  # Try top 3 faculties with lowest load
        faculty = f_data["faculty"]
        
        for room in valid_rooms[:3]:  # Try a few valid rooms
            for day in DAYS:
                for start_t, end_t in TIMESLOTS:
                    # In-memory checks against the rows loaded above. This loop
                    # runs 3 faculty x 3 rooms x 6 days x 6 slots = 324 times,
                    # and each iteration used to issue two queries -- up to 648
                    # round-trips for one request. The early return at five
                    # suggestions only helped when slots were free, so the full
                    # cost landed on a full timetable, which is exactly when the
                    # feature is used. Same comparisons, no round-trips.
                    overlap = any(
                        s.day_of_week == day
                        and s.start_time == start_t
                        and s.end_time == end_t
                        and (s.room_id == room.id or s.faculty_id == faculty.id)
                        for s in semester_schedules
                    )
                    if overlap:
                        continue

                    # Check faculty unavailability
                    blocked = any(
                        u.faculty_id == faculty.id
                        and u.day_of_week == day
                        and u.start_time < end_t
                        and u.end_time > start_t
                        for u in unavailabilities
                    )
                    if blocked:
                        continue
                        
                    # If we reach here, it's a valid slot
                    # Calculate confidence based on faculty load (lower load = higher confidence, min 70%)
                    confidence = int(100 - (f_data["load_percentage"] * 30))
                    
                    suggestions.append({
                        "faculty_id": faculty.id,
                        "faculty_name": f"{faculty.first_name} {faculty.last_name}",
                        "room_id": room.id,
                        "room_name": room.name,
                        "day_of_week": day,
                        "start_time": start_t.strftime("%H:%M"),
                        "end_time": end_t.strftime("%H:%M"),
                        "confidence": confidence
                    })
                    
                    if len(suggestions) >= 5:
                        # Sort suggestions by confidence descending
                        suggestions.sort(key=lambda x: x["confidence"], reverse=True)
                        return suggestions

    # Sort suggestions by confidence descending
    suggestions.sort(key=lambda x: x["confidence"], reverse=True)
    return suggestions

@router.get("/export/pdf")
def export_pdf(
    semester_id: int,
    faculty_name: Optional[str] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Export the current department's schedule for a semester to PDF.

    Columns and row order match the export preview on the Schedule screen, and
    `faculty_name` applies the same optional single-professor scope, so the file
    someone downloads is the one they were shown. Landscape, because nine
    columns do not fit portrait without shrinking the type past reading size.
    """
    if current_user.role not in ['admin', 'program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")

    # Fetch data
    query = db.query(models.Schedule).filter(models.Schedule.semester_id == semester_id)

    dept = None
    dept_name = "All Departments"
    if current_user.role in ['program_chair', 'coordinator']:
        dept = db.query(models.Department).filter(
            (models.Department.code == current_user.department) |
            (models.Department.name == current_user.department)
        ).first()
        if dept:
            query = query.join(models.Curriculum).filter(models.Curriculum.department_id == dept.id)
            dept_name = dept.name

    # Eager-load the three relationships the row loop reads. Resolving them
    # per row issued three queries per schedule -- a 500-class export was about
    # 1,500 round-trips, seconds of latency on a networked database.
    schedules = query.options(
        joinedload(models.Schedule.curriculum),
        joinedload(models.Schedule.faculty),
        joinedload(models.Schedule.room),
    ).all()

    semester = db.query(models.Semester).filter(models.Semester.id == semester_id).first()
    term_label = f"{semester.academic_year} {semester.term}" if semester else ""

    # Resolve each row once, then sort the way the preview does: professor,
    # then the week in order, then time of day.
    day_order = {d: i for i, d in enumerate(['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat'])}
    rows = []
    for s in schedules:
        curriculum_item = s.curriculum
        faculty = s.faculty
        room = s.room

        prof = f"{faculty.first_name} {faculty.last_name}" if faculty else "TBA"
        if faculty_name and prof != faculty_name:
            continue

        rows.append({
            "professor": prof,
            "code": curriculum_item.code if curriculum_item else "N/A",
            "name": curriculum_item.name if curriculum_item else "N/A",
            "day": str(s.day_of_week or ""),
            "start": s.start_time.strftime('%I:%M %p').lstrip('0') if s.start_time else "",
            "end": s.end_time.strftime('%I:%M %p').lstrip('0') if s.end_time else "",
            "room": room.name if room else "No room assigned",
            "building": room.building if room else "",
            "section": s.section or "",
            "_sort": (prof, day_order.get(str(s.day_of_week or "")[:3], 9),
                      s.start_time.strftime('%H:%M') if s.start_time else ""),
        })
    rows.sort(key=lambda r: r["_sort"])

    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        leftMargin=24, rightMargin=24, topMargin=28, bottomMargin=28,
    )
    elements = []
    styles = getSampleStyleSheet()
    cell = ParagraphStyle('cell', parent=styles['Normal'], fontSize=7.5, leading=9.5)

    def cell_text(value):
        """
        reportlab parses a Paragraph's text as mini-HTML, so any '<' or '&' in
        real data is markup. A subject named 'Data Structures #2 <Lab>' lost the
        '<Lab>' silently as an unknown tag, and a stray '<' can abort the whole
        render. Escaping first means the file says what the database says.
        """
        return xml_escape(str(value if value is not None else ''))

    scope_line = faculty_name if faculty_name else "All professors"
    elements.append(Paragraph(cell_text(f"ATLAS Schedule — {dept_name}"), styles['Title']))
    elements.append(Paragraph(cell_text(
        f"{term_label} · {scope_line} · {len(rows)} classes · "
        f"generated {datetime.now().strftime('%d %b %Y %H:%M')}"
    ), styles['Normal']))
    elements.append(Paragraph("<br/>", styles['Normal']))

    header = ["Professor", "Subject Code", "Subject Name", "Day",
              "Start Time", "End Time", "Room", "Building", "Section"]
    data = [header]
    for r in rows:
        data.append([
            Paragraph(cell_text(r["professor"]), cell),
            Paragraph(cell_text(r["code"]), cell),
            Paragraph(cell_text(r["name"]), cell),
            Paragraph(cell_text(r["day"]), cell),
            Paragraph(cell_text(r["start"]), cell),
            Paragraph(cell_text(r["end"]), cell),
            Paragraph(cell_text(r["room"]), cell),
            Paragraph(cell_text(r["building"]), cell),
            Paragraph(cell_text(r["section"]), cell),
        ])

    if not rows:
        elements.append(Paragraph("No classes are scheduled for this selection.", styles['Normal']))
    else:
        t = Table(data, repeatRows=1,
                  colWidths=[95, 62, 175, 55, 58, 58, 82, 72, 45])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b26')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7.5),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f5')]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#c9d2cd')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)

    doc.build(elements)
    buffer.seek(0)

    log_activity(db, current_user.id, "Export PDF", f"Exported schedule for {dept_name} to PDF ({len(rows)} classes)", department_id=dept.id if dept else None) # type: ignore

    return StreamingResponse(buffer, media_type="application/pdf",
                             headers=_attachment_header(f"schedule_{dept_name}", "pdf"))

@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    semester_id: int = 1,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Import schedules from an Excel file.
    Expects columns: CurriculumCode, FacultyEmail, RoomName, Day, StartTime, EndTime, Section
    """
    if current_user.role not in ['admin', 'program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
        
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    sheet = wb.active
    if sheet is None:
        raise HTTPException(status_code=400, detail="Excel sheet not found")
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")
    
    # Basic validation
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    required_cols = ['CurriculumCode', 'FacultyEmail', 'RoomName', 'Day', 'StartTime', 'EndTime', 'Section']
    col_indices = {}
    for col in required_cols:
        if col not in header:
            raise HTTPException(status_code=400, detail=f"Missing required column: {col}")
        col_indices[col] = header.index(col)
            
    success_count = 0
    errors = []

    # Resolve the three lookup tables once instead of three queries per row.
    # An import of 500 classes was 1,500 round-trips, all of them repeats: the
    # same handful of curricula, faculty and rooms fetched over and over.
    # Keyed exactly as the per-row queries were, so unmatched rows still fail
    # the same way and produce the same error message.
    # First match wins, matching the `.first()` these replace. `rooms.name` has
    # no unique constraint -- two buildings may both have a "101" -- so a dict
    # comprehension, which keeps the LAST duplicate, would have quietly bound
    # imported rows to a different room than before.
    def _index(rows_, key):
        out = {}
        for r in rows_:
            k = getattr(r, key)
            if k and k not in out:
                out[k] = r
        return out

    curriculum_by_code = _index(db.query(models.Curriculum).order_by(models.Curriculum.id).all(), "code")
    faculty_by_email = _index(db.query(models.Faculty).order_by(models.Faculty.id).all(), "email")
    room_by_name = _index(db.query(models.Room).order_by(models.Room.id).all(), "name")

    for index, row in enumerate(rows[1:], start=2):
        if not row or all(v is None for v in row):
            continue
        try:
            curr_code = str(row[col_indices['CurriculumCode']]).strip() if row[col_indices['CurriculumCode']] is not None else ""
            fac_email = str(row[col_indices['FacultyEmail']]).strip() if row[col_indices['FacultyEmail']] is not None else ""
            rm_name = str(row[col_indices['RoomName']]).strip() if row[col_indices['RoomName']] is not None else ""
            day_val = str(row[col_indices['Day']]).strip() if row[col_indices['Day']] is not None else ""
            start_val = row[col_indices['StartTime']]
            end_val = row[col_indices['EndTime']]
            sec_val = str(row[col_indices['Section']]).strip() if row[col_indices['Section']] is not None else ""

            curriculum_item = curriculum_by_code.get(curr_code)
            faculty = faculty_by_email.get(fac_email)
            room = room_by_name.get(rm_name)
            
            if not curriculum_item or not faculty or not room:
                errors.append(f"Row {index}: Entity not found (Curriculum: {curriculum_item is not None}, Faculty: {faculty is not None}, Room: {room is not None})")
                continue
                
            def parse_time_val(t_val):
                if isinstance(t_val, datetime):
                    return t_val.time()
                if hasattr(t_val, 'hour'):
                    return t_val
                if isinstance(t_val, str):
                    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M%p"):
                        try:
                            return datetime.strptime(t_val.strip(), fmt).time()
                        except ValueError:
                            pass
                return datetime.strptime(str(t_val).strip(), "%H:%M:%S").time()

            start_t = parse_time_val(start_val)
            end_t = parse_time_val(end_val)
            
            new_sched = models.Schedule(
                semester_id=semester_id,
                curriculum_id=curriculum_item.id,
                faculty_id=faculty.id,
                room_id=room.id,
                day_of_week=day_val,
                start_time=start_t,
                end_time=end_t,
                section=sec_val,
                status='draft'
            )
            db.add(new_sched)
            success_count += 1
        except Exception as e:
            errors.append(f"Row {index}: {str(e)}")
            
    db.commit()
    
    dept = db.query(models.Department).filter(
        (models.Department.code == current_user.department) | 
        (models.Department.name == current_user.department)
    ).first()
    log_activity(db, current_user.id, "Import Excel", f"Imported {success_count} schedules from Excel. Errors: {len(errors)}", department_id=dept.id if dept else None) # type: ignore
    
    return {
        "message": f"Successfully imported {success_count} schedules",
        "errors": errors
    }
 
@router.get("/export/excel")
def export_excel(
    semester_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    """
    Export the current department's schedule for a semester to Excel.
    """
    if current_user.role not in ['admin', 'program_chair', 'coordinator']:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
 
    # Fetch data
    query = db.query(models.Schedule).filter(models.Schedule.semester_id == semester_id)
    
    dept_name = "All Departments"
    if current_user.role in ['program_chair', 'coordinator']:
        dept = db.query(models.Department).filter(
            (models.Department.code == current_user.department) | 
            (models.Department.name == current_user.department)
        ).first()
        if dept:
            query = query.join(models.Curriculum).filter(models.Curriculum.department_id == dept.id)
            dept_name = dept.name
            
    # Eager-load the three relationships the row loop reads. Resolving them
    # per row issued three queries per schedule -- a 500-class export was about
    # 1,500 round-trips, seconds of latency on a networked database.
    schedules = query.options(
        joinedload(models.Schedule.curriculum),
        joinedload(models.Schedule.faculty),
        joinedload(models.Schedule.room),
    ).all()

    # Prepare data
    export_data = []
    headers = ["Curriculum Code", "Subject Name", "Section", "Faculty", "Room", "Day", "Start Time", "End Time", "Status", "Locked"]
    for s in schedules:
        curriculum_item = s.curriculum
        faculty = s.faculty
        room = s.room
        
        export_data.append({
            "Curriculum Code": curriculum_item.code if curriculum_item else "N/A",
            "Subject Name": curriculum_item.name if curriculum_item else "N/A",
            "Section": s.section,
            "Faculty": f"{faculty.first_name} {faculty.last_name}" if faculty else "TBA",
            "Room": room.name if room else "N/A",
            "Day": s.day_of_week,
            "Start Time": s.start_time.strftime('%I:%M %p') if s.start_time else "",
            "End Time": s.end_time.strftime('%I:%M %p') if s.end_time else "",
            "Status": s.status,
            "Locked": "Yes" if s.is_locked else "No"
        })
    
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Schedules")
    ws.title = "Schedules"
    ws.append(headers)
    
    for item in export_data:
        ws.append([item[h] for h in headers])
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_activity(db, current_user.id, "Export Excel", f"Exported schedule for {dept_name} to Excel", department_id=dept.id if current_user.role in ['program_chair', 'coordinator'] and dept else None) # type: ignore
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            **_attachment_header(f"schedule_{dept_name}", "xlsx")
        }
    )

# --- Catch-all parameter route: must stay last so it cannot shadow the
# literal paths declared above (/suggestions, /export/*, /restore, /clear-all).

@router.get("/{schedule_id}", response_model=schemas.ScheduleResponse)
def get_schedule(
    schedule_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    schedule = db.query(models.Schedule).options(
        joinedload(models.Schedule.curriculum),
        joinedload(models.Schedule.room)
    ).filter(models.Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Schedule not found")

    if current_user.role in ['program_chair', 'coordinator']:
        curriculum_item = db.query(models.Curriculum).filter(models.Curriculum.id == schedule.curriculum_id).first()
        if curriculum_item:
            dept = db.query(models.Department).filter(models.Department.id == curriculum_item.department_id).first()
            if not dept or (dept.code != current_user.department and dept.name != current_user.department):
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")

    return schedule
