from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile, Form, Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
import openpyxl
import math
import io
import json
import re
from .. import models, schemas, database, auth
from .logs import log_activity

router = APIRouter(
    prefix="/api/curriculum",
    tags=["Curriculum"]
)

@router.get("/blocks", response_model=List[schemas.CurriculumBlockWithCount])
def get_curriculum_blocks(
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    try:
        query = db.query(models.CurriculumBlock)
        
        if current_user.role != 'admin':
            # All non-admin roles only see PUBLISHED curricula
            query = query.filter(models.CurriculumBlock.status == 'PUBLISHED')

        if current_user.role in ['program_chair', 'coordinator']:
            dept = db.query(models.Department).filter(
                (models.Department.code == current_user.department) | 
                (models.Department.name == current_user.department)
            ).first()
            if dept:
                query = query.filter(models.CurriculumBlock.department_id == dept.id)
            else:
                return []

        blocks = query.all()
        results = []
        
        for block in blocks:
            subjects = db.query(models.Curriculum).filter(models.Curriculum.block_id == block.id).all()
            results.append({
                "id": block.id,
                "program_name": block.program_name,
                "academic_year": block.academic_year,
                "filename": block.filename,
                "department_id": block.department_id,
                "program_id": getattr(block, "program_id", None),
                "status": getattr(block, "status", "PUBLISHED") or "PUBLISHED",
                "created_at": block.created_at,
                "subject_count": len(subjects),
                "total_units": sum(s.units for s in subjects)
            })
            
        return results
    except Exception as e:
        print(f"Fallback handling status query error in get_curriculum_blocks: {e}")
        blocks = db.query(models.CurriculumBlock).all()
        results = []
        for block in blocks:
            b_status = getattr(block, "status", "PUBLISHED") or "PUBLISHED"
            if current_user.role != 'admin' and b_status != 'PUBLISHED':
                continue
            if current_user.role in ['program_chair', 'coordinator']:
                dept = db.query(models.Department).filter(
                    (models.Department.code == current_user.department) | 
                    (models.Department.name == current_user.department)
                ).first()
                if dept and block.department_id != dept.id:
                    continue

            subjects = db.query(models.Curriculum).filter(models.Curriculum.block_id == block.id).all()
            results.append({
                "id": block.id,
                "program_name": block.program_name,
                "academic_year": block.academic_year,
                "filename": block.filename,
                "department_id": block.department_id,
                "program_id": getattr(block, "program_id", None),
                "status": b_status,
                "created_at": block.created_at,
                "subject_count": len(subjects),
                "total_units": sum(s.units for s in subjects)
            })
        return results

@router.post("/blocks", response_model=schemas.CurriculumBlockWithCount, status_code=status.HTTP_201_CREATED)
def create_curriculum_block(
    block_data: schemas.CurriculumBlockCreate,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only System Administrators can create curriculum flowcharts")
        
    target_dept_id = block_data.department_id

    # A version belongs to a programme; the college follows from it. Deriving
    # the college here stops a block being filed under one college while its
    # programme sits in another.
    if block_data.program_id:
        program = db.query(models.Program).filter(models.Program.id == block_data.program_id).first()
        if not program:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Programme not found.")
        target_dept_id = program.department_id

    existing = db.query(models.CurriculumBlock).filter(
        models.CurriculumBlock.program_name == block_data.program_name,
        models.CurriculumBlock.academic_year == block_data.academic_year,
        models.CurriculumBlock.department_id == target_dept_id
    ).first()
    if existing:
        subjects = db.query(models.Curriculum).filter(models.Curriculum.block_id == existing.id).all()
        return {
            "id": existing.id,
            "program_name": existing.program_name,
            "academic_year": existing.academic_year,
            "filename": existing.filename,
            "department_id": existing.department_id,
            "program_id": getattr(existing, "program_id", None),
            "status": getattr(existing, "status", "PUBLISHED") or "PUBLISHED",
            "created_at": existing.created_at,
            "subject_count": len(subjects),
            "total_units": sum(s.units for s in subjects)
        }

    new_block = models.CurriculumBlock(
        program_name=block_data.program_name.strip(),
        academic_year=block_data.academic_year.strip(),
        filename=block_data.filename or "Manual Entry",
        department_id=target_dept_id,
        program_id=block_data.program_id,
        status=block_data.status or 'PUBLISHED'
    )
    db.add(new_block)
    db.commit()
    db.refresh(new_block)
    
    log_activity(db, current_user.id, "Create Curriculum Block", f"Created curriculum block: {new_block.program_name} ({new_block.academic_year})", "success", department_id=target_dept_id) # type: ignore

    return {
        "id": new_block.id,
        "program_name": new_block.program_name,
        "academic_year": new_block.academic_year,
        "filename": new_block.filename,
        "department_id": new_block.department_id,
        "program_id": getattr(new_block, "program_id", None),
        "status": getattr(new_block, "status", "PUBLISHED") or "PUBLISHED",
        "created_at": new_block.created_at,
        "subject_count": 0,
        "total_units": 0
    }

@router.patch("/blocks/{block_id}/status", response_model=schemas.CurriculumBlockWithCount)
def update_curriculum_block_status(
    block_id: int,
    status_value: str = Form(..., alias="status"),
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only System Administrators can change curriculum status")
        
    block = db.query(models.CurriculumBlock).filter(models.CurriculumBlock.id == block_id).first()
    if not block:
        raise HTTPException(status_code=404, detail="Curriculum block not found")
        
    new_status = status_value.upper().strip()
    if new_status not in ['DRAFT', 'PUBLISHED', 'ARCHIVED']:
        raise HTTPException(status_code=400, detail="Invalid status. Must be DRAFT, PUBLISHED, or ARCHIVED.")
        
    setattr(block, "status", new_status)
    db.commit()
    db.refresh(block)
    
    log_activity(db, current_user.id, "Update Curriculum Status", f"Updated curriculum {block.program_name} status to {new_status}", "success", department_id=block.department_id) # type: ignore
    
    subjects = db.query(models.Curriculum).filter(models.Curriculum.block_id == block.id).all()
    return {
        "id": block.id,
        "program_name": block.program_name,
        "academic_year": block.academic_year,
        "filename": block.filename,
        "department_id": block.department_id,
        "status": block.status,
        "created_at": block.created_at,
        "subject_count": len(subjects),
        "total_units": sum(s.units for s in subjects)
    }

@router.get("", response_model=List[schemas.CurriculumResponse])
def get_curriculum(
    skip: int = 0,
    # Was 100. A single curriculum version can hold more than that — DVM has
    # 117 subjects — so the default quietly truncated a whole programme.
    limit: int = 1000,
    department_id: Optional[int] = None,
    program_code: Optional[str] = None,
    block_id: Optional[int] = None,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    query = db.query(models.Curriculum)
    
    if current_user.role in ['program_chair', 'coordinator'] and current_user.department:
        # Filter curriculum by the program chair/coordinator's department
        dept = db.query(models.Department).filter(
            (models.Department.code == current_user.department) | 
            (models.Department.name == current_user.department)
        ).first()
        if dept:
            query = query.filter(models.Curriculum.department_id == dept.id)
        else:
            return [] # No department match means no access
    elif department_id:
        query = query.filter(models.Curriculum.department_id == department_id)

    if program_code:
        query = query.filter(models.Curriculum.program_code == program_code)
    
    if block_id:
        query = query.filter(models.Curriculum.block_id == block_id)
        
    return query.offset(skip).limit(limit).all()

@router.get("/{curriculum_id}", response_model=schemas.CurriculumResponse)
def get_curriculum_item(
    curriculum_id: int, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    curriculum_item = db.query(models.Curriculum).filter(models.Curriculum.id == curriculum_id).first()
    if not curriculum_item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Curriculum item not found")
        
    if current_user.role in ['program_chair', 'coordinator'] and current_user.department:
        dept = db.query(models.Department).filter(models.Department.id == curriculum_item.department_id).first()
        if not dept or (dept.code != current_user.department and dept.name != current_user.department):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized to access this curriculum item")
            
    return curriculum_item

@router.post("", response_model=schemas.CurriculumResponse, status_code=status.HTTP_201_CREATED)
def create_curriculum_item(
    curriculum_item: schemas.CurriculumCreate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only System Administrators can create curriculum subjects")
        
    if not curriculum_item.department_id:
        first_dept = db.query(models.Department).first()
        if first_dept:
            curriculum_item.department_id = int(first_dept.id) # type: ignore
             
    # Determine subject type using Lec/Lab unit values as primary priority
    lec_u_val = curriculum_item.lec_units
    lab_u_val = curriculum_item.lab_units
    if lec_u_val > 0 and lab_u_val == 0:
        curriculum_item.type = 'lecture'
    elif lec_u_val == 0 and lab_u_val > 0:
        curriculum_item.type = 'lab'

    ab_match = re.search(r"^(.*?)[_\-\s]*A/B$", curriculum_item.code, re.IGNORECASE)
    # Split only when both halves are real; see the import path for why an
    # A/B code with no laboratory units must not gain a fabricated one.
    if ab_match and curriculum_item.lab_units > 0 and curriculum_item.lec_units > 0:
        base_code = ab_match.group(1).strip()
        code_a = f"{base_code}A"
        code_b = f"{base_code}B"

        u_lec = curriculum_item.lec_units
        u_lab = curriculum_item.lab_units

        existing_a = db.query(models.Curriculum).filter(
            models.Curriculum.block_id == curriculum_item.block_id,
            func.lower(models.Curriculum.code) == code_a.lower()
        ).first() if curriculum_item.block_id else None

        existing_b = db.query(models.Curriculum).filter(
            models.Curriculum.block_id == curriculum_item.block_id,
            func.lower(models.Curriculum.code) == code_b.lower()
        ).first() if curriculum_item.block_id else None

        if existing_a and existing_b:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Both {code_a} and {code_b} already exist in this curriculum block")

        created_item = None
        if not existing_a:
            new_a = models.Curriculum(
                **{**curriculum_item.model_dump(), "code": code_a, "type": "lecture", "units": u_lec, "lec_units": u_lec, "lab_units": 0}
            )
            db.add(new_a)
            created_item = new_a
        if not existing_b:
            new_b = models.Curriculum(
                **{**curriculum_item.model_dump(), "code": code_b, "type": "lab", "units": u_lab, "lec_units": 0, "lab_units": u_lab}
            )
            db.add(new_b)
            if not created_item: created_item = new_b

        db.commit()
        if created_item: db.refresh(created_item)
        log_activity(db, current_user.id, "Create Curriculum", f"Created split A/B subjects: {code_a}, {code_b}", "success", department_id=curriculum_item.department_id) # type: ignore
        return created_item or existing_a

    # Check per-block duplicate
    if curriculum_item.block_id:
        existing_single = db.query(models.Curriculum).filter(
            models.Curriculum.block_id == curriculum_item.block_id,
            func.lower(models.Curriculum.code) == curriculum_item.code.lower()
        ).first()
        if existing_single:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Subject {curriculum_item.code} already exists in this curriculum block")
    else:
        db_curriculum = db.query(models.Curriculum).filter(
            models.Curriculum.code == curriculum_item.code,
            models.Curriculum.program_code == curriculum_item.program_code
        ).first()
        if db_curriculum:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Curriculum item with this code already exists for this program")
        
    new_curriculum = models.Curriculum(**curriculum_item.model_dump())
    db.add(new_curriculum)
    db.commit()
    db.refresh(new_curriculum)
    
    log_activity(db, current_user.id, "Create Curriculum", f"Created subject: {new_curriculum.code}", "success", department_id=new_curriculum.department_id) # type: ignore
    
    return new_curriculum

@router.put("/{curriculum_id}", response_model=schemas.CurriculumResponse)
def update_curriculum_item(
    curriculum_id: int, 
    curriculum_item: schemas.CurriculumUpdate, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only System Administrators can edit curriculum subjects")
        
    db_curriculum = db.query(models.Curriculum).filter(models.Curriculum.id == curriculum_id).first()
    if not db_curriculum:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Curriculum item not found")
             
    update_data = curriculum_item.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_curriculum, key, value)
        
    db.commit()
    db.refresh(db_curriculum)
    
    log_activity(db, current_user.id, "Update Curriculum", f"Updated subject: {db_curriculum.code}", "success", department_id=db_curriculum.department_id) # type: ignore
    
    return db_curriculum

@router.delete("/{curriculum_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_curriculum_item(
    curriculum_id: int, 
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only System Administrators can delete curriculum subjects")
        
    db_curriculum = db.query(models.Curriculum).filter(models.Curriculum.id == curriculum_id).first()
    if not db_curriculum:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Curriculum item not found")
             
    curr_code = db_curriculum.code
    dept_id = getattr(db_curriculum, 'department_id', None)
    db.delete(db_curriculum)
    db.commit()
    
    log_activity(db, current_user.id, "Delete Curriculum", f"Deleted subject: {curr_code}", "success", department_id=dept_id) # type: ignore
    
    return None

async def _process_curriculum_import(
    contents: bytes,
    department_id: Optional[int],
    program_code: Optional[str],
    dry_run: bool,
    db: Session,
    current_user: models.User,
    custom_mapping: Optional[str] = None,
    selected_sheet: Optional[str] = None
):
    # Import is the last curriculum write a non-admin could reach. Every other
    # handler in this router is already admin-only, but this one accepted chairs
    # and coordinators -- and with dry_run=false it creates a CurriculumBlock and
    # its subjects outright, so a chair could author curriculum the catalog was
    # meant to hand them. The curriculum is institutional reference data: the
    # administrator owns it, departments read it.
    if current_user.role != 'admin':
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only System Administrators can import curriculum Excel files"
        )

    # Determine target department
    target_dept_id = department_id
    if not target_dept_id:
        first_dept = db.query(models.Department).first()
        if not first_dept:
             raise HTTPException(status_code=400, detail="No departments found in system")
        target_dept_id = first_dept.id

    try:
        import openpyxl, math
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        
        items_to_add = []
        skipped_items = []
        all_parsed_data = [] 
        
        mapping_keywords = {
            'code': ["course code", "code", "subject code", "catalog"],
            'name': ["course title", "title", "subject name", "description", "subject"],
            'units': ["units", "unit", "credit", "total units"],
            'lec_units': ["lec", "lecture", "lec units", "lec. units"],
            'lab_units': ["lab", "laboratory", "lab units", "lab. units"],
            'pre_requisite': ["pre-req", "prerequisite", "prerequisite(s)", "pre-requisite", "prereq"],
            'year_level': ["year", "yr", "grade"],
            'semester_term': ["sem", "semester", "term"]
        }

        if custom_mapping:
            try:
                user_map = json.loads(custom_mapping)
                for k, v in user_map.items():
                    if k in mapping_keywords:
                        mapping_keywords[k] = [v] if isinstance(v, str) else v
            except:
                pass

        best_sheet = None
        max_score = -1
        best_ay = 0
        extracted_program_name = "Unknown Program"
        extracted_ay = "Unknown AY"
        
        program_patterns = [
            r'BACHELOR OF SCIENCE IN [A-Z\s]+',
            r'BACHELOR OF [A-Z\s]+',
            r'DOCTOR OF [A-Z\s]+',
            r'MASTER OF [A-Z\s]+',
            r'ASSOCIATE IN [A-Z\s]+',
            r'\bBS[A-Z0-9\-\s]{1,10}\b',
            r'\bAB[A-Z0-9\-\s]{1,10}\b',
            r'\bDVM\b'
        ]

        def is_na(v):
            if v is None: return True
            if isinstance(v, float) and math.isnan(v): return True
            return False

        # Year/term section headers are what a full curriculum grid is made of,
        # so they are the strongest evidence a sheet holds the real curriculum.
        section_re = re.compile(r'\b(FIRST|SECOND|THIRD|FOURTH|FIFTH|SIXTH)\s+YEAR\b')
        term_re = re.compile(r'\b(FIRST|SECOND|THIRD|SUMMER|MIDYEAR)\s+TERM\b')

        for sheet_name in wb.sheetnames:
            try:
                ws_probe = wb[sheet_name]
                all_rows = [list(r) for r in ws_probe.iter_rows(values_only=True)]
                raw_probe_rows = all_rows[:25]
            except Exception:
                continue

            probe_vals = []
            for r in raw_probe_rows:
                for v in r:
                    if not is_na(v):
                        probe_vals.append(str(v).strip())

            probe_text = " ".join([v.upper() for v in probe_vals])
            probe_vals_lower = [v.lower() for v in probe_vals]

            score = 0
            has_code_hdr = any(any(k == val or (len(k) > 3 and k in val) for k in mapping_keywords['code']) for val in probe_vals_lower)
            has_title_hdr = any(any(k == val or (len(k) > 3 and k in val) for k in mapping_keywords['name']) for val in probe_vals_lower)
            has_unit_hdr = any(any(k == val or (len(k) > 3 and k in val) for k in mapping_keywords['units']) for val in probe_vals_lower)

            if has_code_hdr: score += 40
            if has_title_hdr: score += 30
            if has_unit_hdr: score += 20

            ay_match = re.search(r'AY\s*(\d{4})', probe_text)
            if ay_match:
                score += 10

            for pat in program_patterns:
                if re.search(pat, probe_text):
                    score += 10
                    break

            # Completeness, measured over the whole sheet rather than the first
            # 25 rows.
            #
            # Without this every sheet in a workbook scored identically on
            # header presence alone, and the tie was broken by sheet order — so
            # a workbook carrying ten tabs imported whichever happened to be
            # first. For BSCS that was a superseded AY2018 copy sitting ahead of
            # the current "UPDATED" sheet, and the import silently produced the
            # wrong curriculum while reporting high confidence.
            sections = 0
            subject_rows = 0
            for row in all_rows:
                cells = [str(v).strip() for v in row if not is_na(v)]
                if not cells:
                    continue
                joined = " ".join(cells).upper()
                if section_re.search(joined) and term_re.search(joined):
                    sections += 1
                # A subject row: a code-shaped token plus at least one number.
                elif any(re.match(r'^[A-Z]{2,6}\s?\d{2,4}', c.upper()) for c in cells) and \
                     any(re.fullmatch(r'\d{1,2}(\.\d+)?', c) for c in cells):
                    subject_rows += 1

            score += sections * 25 + min(subject_rows, 200)

            # Most recent effectivity wins a genuine tie: a workbook usually
            # keeps the superseded year alongside the current one.
            ay_year = int(ay_match.group(1)) if ay_match else 0

            candidate = (score, ay_year)
            best = (max_score, best_ay)
            if candidate > best and (has_code_hdr or has_title_hdr):
                max_score, best_ay = candidate
                best_sheet = sheet_name

        if not best_sheet:
            best_sheet = wb.sheetnames[0]

        # Sheet detection is a heuristic, so the caller must be able to overrule
        # it. Commit 7c1bf6d claimed this parameter existed; it did not, which
        # meant a workbook whose real curriculum sat on a lower-scoring tab
        # could not be imported at all.
        sheet_auto_selected = True
        if selected_sheet:
            if selected_sheet not in wb.sheetnames:
                raise HTTPException(
                    status_code=400,
                    detail=f"'{selected_sheet}' is not a sheet in this workbook. Available: {', '.join(wb.sheetnames)}"
                )
            best_sheet = selected_sheet
            sheet_auto_selected = False

        print(f"DEBUG: Selected sheet '{best_sheet}' with confidence score {max_score}")

        ws_best = wb[best_sheet]
        raw_probe_rows = [list(r) for r in list(ws_best.iter_rows(values_only=True))[:25]]
        probe_vals = []
        for r in raw_probe_rows:
            for v in r:
                if not is_na(v):
                    probe_vals.append(str(v).strip())
        probe_text = " ".join([v.upper() for v in probe_vals])

        ay_match = re.search(r'(?:AY|ACADEMIC YEAR)\s*:?\s*(\d{4}\s*[-–]\s*\d{4}|\d{4})', probe_text)
        if ay_match:
            extracted_ay = f"AY {ay_match.group(1).strip()}"
        else:
            ay_simple = re.search(r'\b(20\d{2}\s*[-–]\s*20\d{2})\b', probe_text)
            if ay_simple:
                extracted_ay = f"AY {ay_simple.group(1).strip()}"

        for pat in program_patterns:
            prog_match = re.search(pat, probe_text)
            if prog_match:
                extracted_program_name = prog_match.group(0).strip()
                break

        if extracted_program_name == "Unknown Program":
            clean_sheet_title = best_sheet.replace("_", " ").replace("-", " ").strip()
            if len(clean_sheet_title) > 2 and not clean_sheet_title.lower().startswith("sheet"):
                extracted_program_name = clean_sheet_title.upper()

        words = [w for w in re.split(r'[\s\-]+', extracted_program_name) if w not in ["OF", "IN", "SCIENCE", "BACHELOR", "ASSOCIATE", "DOCTOR", "MASTER", "AND"]]
        if words:
            program_code = "".join([w[0] for w in words])
        else:
            program_code = extracted_program_name[:4].upper()
            
        if len(program_code) < 2:
            program_code = (best_sheet[:4] if len(best_sheet) >= 4 else best_sheet).upper()

        full_block_name = f"{program_code} ({extracted_ay})" if extracted_ay != "Unknown AY" else f"{program_code} (Curriculum)"

        existing_block = db.query(models.CurriculumBlock).filter(
            models.CurriculumBlock.department_id == target_dept_id,
            models.CurriculumBlock.program_name == extracted_program_name,
            models.CurriculumBlock.academic_year == extracted_ay
        ).first()

        if existing_block:
            curriculum_block = existing_block
            print(f"DEBUG: Found existing block '{extracted_program_name} ({extracted_ay})' (ID: {curriculum_block.id})")
        else:
            curriculum_block = models.CurriculumBlock(
                department_id=target_dept_id,
                program_name=extracted_program_name,
                academic_year=extracted_ay,
                filename="Uploaded File"
            )
            db.add(curriculum_block)
            db.flush()
            print(f"DEBUG: Created new block '{full_block_name}' (ID: {curriculum_block.id})")

        ws = wb[best_sheet]
        
        merged_cells = list(ws.merged_cells.ranges)
        for merged_range in merged_cells:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c).value = top_left_value # type: ignore
        
        raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        
        # Helper functions for context normalization
        def normalize_year(value: object) -> Optional[str]:
            if is_na(value):
                return None
            s = str(value).strip().upper()
            if not s or s in {"NAN", "NONE", "N/A"}: return None
            m = re.search(r"\b([1-5])\b", s)
            if m: return m.group(1)
            m = re.search(r"\b(1ST|2ND|3RD|4TH|5TH)\b", s)
            if m: return m.group(1)[0]
            # SIXTH matters: DVM is a six-year programme, and without it the
            # whole sixth year was skipped on import — 33 units of clinical
            # internship silently absent from the curriculum.
            word_map = {"FIRST": "1", "SECOND": "2", "THIRD": "3", "FOURTH": "4", "FIFTH": "5", "SIXTH": "6"}
            for k, v in word_map.items():
                if k in s: return v
            return None

        def normalize_semester(value: object) -> Optional[str]:
            if is_na(value):
                return None
            s = str(value).strip().upper()
            if not s or s in {"NAN", "NONE", "N/A"}: return None
            if "3RD SEMESTER" in s or "MIDYEAR" in s: return "3rd semester"
            m = re.search(r"\b([123])\b", s)
            if m: return {"1": "1st", "2": "2nd", "3": "3rd"}[m.group(1)]
            if re.search(r"\bI\b", s): return "1st"
            if re.search(r"\bII\b", s): return "2nd"
            if re.search(r"\bIII\b", s): return "3rd"
            if "1ST" in s or "FIRST" in s: return "1st"
            if "2ND" in s or "SECOND" in s: return "2nd"
            if "3RD" in s or "THIRD" in s: return "3rd"
            return None
        
        # STEP 7 — DEDUPLICATION & CONTEXT TRACKER
        last_row_identifier = None
        current_year_context = None
        current_sem_context = None
        is_in_active_zone = False
        looking_for_headers = True
        seen_any_section = False
        col_map = {} 

        for idx, row in enumerate(raw_rows):
            row_text = " ".join([str(v).strip().upper() for v in row if not is_na(v)])
            if not row_text: continue

            # 1. Independent Section & Zone Header Detection
            #
            # `row_text` joins every cell, prerequisites included, and a great
            # many subjects carry one like "4th Year Standing" or "3rd Year
            # Standing". Matched blindly, those read as a new Year 4 section and
            # the subject on that row was dropped without a word — BSCS lost its
            # entire Fourth Year (On the Job Training) exactly this way, and a
            # dozen more subjects besides.
            #
            # A section header announces a section and nothing else. A row that
            # carries both a course code and a unit count is a subject, whatever
            # its prerequisite happens to say.
            cell_strs = [str(v).strip() for v in row if not is_na(v)]
            looks_like_subject = (
                any(re.match(r'^[A-Z]{2,6}\s?\d{2,4}', c.upper()) for c in cell_strs)
                and any(re.fullmatch(r'\d{1,2}(\.\d+)?', c) for c in cell_strs)
            )

            y_match = None if looks_like_subject else re.search(r'(1ST|2ND|3RD|4TH|5TH|6TH|FIRST|SECOND|THIRD|FOURTH|FIFTH|SIXTH)\s+YEAR|YEAR\s+([1-6])', row_text)
            s_match = None if looks_like_subject else re.search(r'(1ST|2ND|3RD|FIRST|SECOND|THIRD)\s+(SEMESTER|TERM)|(3RD SEMESTER|MIDYEAR)', row_text)

            section_updated = False
            if y_match:
                y_raw = y_match.group(1) or y_match.group(2)
                norm_y = normalize_year(y_raw)
                if norm_y:
                    current_year_context = norm_y
                    section_updated = True

            if s_match:
                s_raw = s_match.group(1) or s_match.group(3)
                norm_s = normalize_semester(s_raw)
                if norm_s:
                    current_sem_context = norm_s
                    section_updated = True

            if section_updated:
                is_in_active_zone = True
                looking_for_headers = True 
                last_row_identifier = None # Reset for new zone
                seen_any_section = True
                print(f"DEBUG: Found Section Header at row {idx}: Year={current_year_context}, Sem={current_sem_context}")
                continue

            # Professional Electives pool.
            #
            # A curriculum sheet lists these after the year/term grid: subjects
            # a student chooses from, not subjects taught in a given term. The
            # import covers what is scheduled term by term, so the pool closes
            # the active zone rather than opening a pseudo-term. Keeping it out
            # is also what makes each programme's imported total equal the
            # TOTAL UNITS printed on its own sheet.
            if "ELECTIVES" in row_text:
                current_year_context = None
                current_sem_context = None
                is_in_active_zone = False
                looking_for_headers = False
                last_row_identifier = None
                continue

            # Summary of Units Trigger (Close Zone)
            if "SUMMARY OF UNITS" in row_text:
                is_in_active_zone = False
                continue

            # 2. Universal Table Header Detection
            row_vals = [str(v).strip().lower() if not is_na(v) else "" for v in row]
            temp_map = {}
            for key, keywords in mapping_keywords.items():
                for i, val in enumerate(row_vals):
                    if any(k == val or (len(k) > 3 and k in val) for k in keywords):
                        if key == 'units' and any(k in val for k in ["lec", "lab"]): continue
                        temp_map[key] = i
                        break

            if 'code' in temp_map:
                col_map = temp_map
                looking_for_headers = False
                is_in_active_zone = True
                # Only assume Year 1 / 1st Term for a sheet that has no year or
                # term structure at all. Applying it after a section has been
                # seen would refile whatever follows the last term -- the
                # electives pool, most obviously -- into First Year.
                if not seen_any_section:
                    if not current_year_context:
                        current_year_context = '1'
                    if not current_sem_context:
                        current_sem_context = '1st'
                print(f"DEBUG: Mapped columns at row {idx}: {col_map}")
                continue

            # 3. Subject Row Capture
            #
            # A subject is imported only when it sits inside a year AND a term.
            # Everything a curriculum sheet carries outside that grid --
            # electives pools, unit summaries, signatories -- is not something
            # that gets scheduled, so it is not curriculum for this purpose.
            if not is_in_active_zone or not col_map or 'code' not in col_map: continue
            if not current_year_context or not current_sem_context: continue
            if not str(current_year_context).isdigit(): continue
            
            code_raw = row[col_map['code']] if col_map['code'] < len(row) else None
            name_raw = row[col_map['name']] if 'name' in col_map and col_map['name'] < len(row) else ""
            units_raw = row[col_map['units']] if 'units' in col_map and col_map['units'] < len(row) else 0
            
            def parse_num(val):
                if is_na(val): return 0
                s = re.sub(r'[\(\)\[\]\*\s]', '', str(val).strip())
                try: return int(float(s))
                except: return 0
            
            units = parse_num(units_raw)
            
            if is_na(code_raw):
                if units > 0:
                    print(f"DEBUG: Row {idx} skipped: Subtotal row (units={units}, code=None)")
                    continue 
                if not is_na(name_raw) and current_year_context == 'Elective':
                    sub_cat = str(name_raw).strip()
                    if sub_cat and len(sub_cat) > 3:
                        current_sem_context = sub_cat
                        print(f"DEBUG: Found Elective Sub-category at row {idx}: {sub_cat}")
                continue
            
            # Cleanup code and name for identifier
            code = str(int(code_raw)) if isinstance(code_raw, float) and code_raw.is_integer() else str(code_raw).strip()
            name = str(name_raw).strip() if not is_na(name_raw) else ""
            print(f"DEBUG: Found Subject candidate at row {idx}: {code} - {name} ({units} units)")
            # STEP 7 — CONSECUTIVE DEDUPLICATION (Scenario A, Rule 260)
            current_identifier = (code, name, units)
            if current_identifier == last_row_identifier:
                continue # Skip merged cell artifact
            last_row_identifier = current_identifier
            blacklist = [
                'prepared:', 'noted:', 'approved:', 'grade', 'course code', 'course title', 
                'total units', 'summary of units', 'signature', 'printed name', 'page ', 'rev.'
            ]
            if not code or any(b in code.lower() for b in blacklist) or len(code) < 2:
                continue
            
            lec_units = parse_num(row[col_map['lec_units']] if 'lec_units' in col_map and col_map['lec_units'] < len(row) else 0)
            lab_units = parse_num(row[col_map['lab_units']] if 'lab_units' in col_map and col_map['lab_units'] < len(row) else 0)
            
            if units == 0 and (lec_units > 0 or lab_units > 0):
                units = lec_units + lab_units
            
            # Rule 205: Must have non-zero Units
            if units == 0: continue

            if any(k in name.lower() for k in ["course title", "subtotal", "total units", "description"]):
                continue

            def clean_prereqs(val):
                if not val or is_na(val): return None
                s = str(val).lower().strip()
                if s in ['none', 'n/a', '0', 'nan', 'none.', 'no']: return None
                parts = str(val).replace('\r\n', ',').replace('\n', ',').replace('&', ',').replace(';', ',').replace(' AND ', ',').replace(' OR ', ',')
                raw_tokens = [c.strip().upper() for c in parts.split(',') if c and c.strip()]
                codes: list[str] = []
                for tok in raw_tokens:
                    tok_clean = re.sub(r"\s+", " ", tok).strip()
                    if tok_clean in {"AND", "OR", "NONE", "N/A", "0"}: continue
                    if tok_clean:
                        codes.append(tok_clean)
                return ", ".join(codes) if codes else None

            pre_req_raw = row[col_map['pre_requisite']] if 'pre_requisite' in col_map and col_map['pre_requisite'] < len(row) else None
            pre_req = clean_prereqs(pre_req_raw)
            is_major_val = not any(code.upper().strip().startswith(prefix) for prefix in ('CORE', 'PEED', 'NSTP', 'LSVI', 'GE', 'RZAL', 'RIZAL'))

            # DYNAMIC A/B SUBJECT SPLITTING ([BASE SUBJECT CODE]A/B -> [BASE SUBJECT CODE]A + [BASE SUBJECT CODE]B)
            ab_match = re.search(r"^(.*?)[_\-\s]*A/B$", code, re.IGNORECASE)
            # Split only when the row really carries both halves. Some subjects
            # are written "IAS101A/B" but list Lec 2 / Lab 0 / Units 2 — a
            # lecture-only subject with a combined-style code. Splitting those
            # invented a 1-unit laboratory that exists nowhere in the workbook
            # and pushed the term's total above the printed subtotal.
            if ab_match and lab_units > 0 and lec_units > 0:
                base_code = ab_match.group(1).strip()
                code_a = f"{base_code}A"
                code_b = f"{base_code}B"

                u_lec = lec_units
                u_lab = lab_units

                split_specs = [
                    (code_a, "lecture", u_lec, u_lec, 0),
                    (code_b, "lab", u_lab, 0, u_lab)
                ]

                for ccode, ctype_val, u_val, lec_u, lab_u in split_specs:
                    item_data = {
                        "block_id": curriculum_block.id,
                        "code": ccode, "name": name, "units": u_val, "type": ctype_val,
                        "department_id": target_dept_id, "program_code": program_code,
                        "year_level": current_year_context, "semester_term": current_sem_context,
                        "lec_units": lec_u, "lab_units": lab_u, "pre_requisite": pre_req,
                        "is_major": is_major_val,
                        "validation_issues": []
                    }

                    is_duplicate_in_run = any(i.code == ccode for i in items_to_add)
                    if is_duplicate_in_run:
                        item_data["validation_issues"].append("Duplicate: Already exists in this file")
                        skipped_items.append({**item_data, "reason": "Already exists in this file"})
                    else:
                        items_to_add.append(models.Curriculum(
                            block_id=curriculum_block.id,
                            code=ccode, name=name, units=u_val, type=ctype_val,
                            department_id=target_dept_id, program_code=program_code,
                            year_level=current_year_context, semester_term=current_sem_context,
                            lec_units=lec_u, lab_units=lab_u, pre_requisite=pre_req,
                            is_major=is_major_val
                        ))
                    all_parsed_data.append(item_data)
                    print(f"DEBUG: Captured split A/B subject: {ccode} ({ctype_val})")
                continue

            # Determine subject type using Lec and Lab columns as primary source of truth:
            if lec_units > 0 and lab_units == 0:
                ctype = 'lecture'
            elif lec_units == 0 and lab_units > 0:
                ctype = 'lab'
            else:
                if code.upper().endswith('B') or 'lab' in name.lower() or 'laboratory' in name.lower():
                    ctype = 'lab'
                else:
                    ctype = 'lecture'

            item_data = {
                "block_id": curriculum_block.id,
                "code": code, "name": name, "units": units, "type": ctype,
                "department_id": target_dept_id, "program_code": program_code,
                "year_level": current_year_context, "semester_term": current_sem_context,
                "lec_units": lec_units, "lab_units": lab_units, "pre_requisite": pre_req,
                "is_major": is_major_val,
                "validation_issues": []
            }
            
            # Unit Mismatch Warning check
            if (lec_units > 0 or lab_units > 0) and units != (lec_units + lab_units):
                item_data["validation_issues"].append(
                    f"Unit Mismatch: Total ({units}) != Lec ({lec_units}) + Lab ({lab_units})"
                )

            if not current_year_context or not current_sem_context:
                item_data["validation_issues"].append(
                    "WARNING: Year/Semester context could not be determined"
                )

            # Duplication check WITHIN the block (Step 0 rule: Isolated)
            is_duplicate_in_run = any(i.code == code for i in items_to_add)

            if is_duplicate_in_run:
                item_data["validation_issues"].append("Duplicate: Already exists in this file")
                skipped_items.append({**item_data, "reason": "Already exists in this file"})
            else:
                items_to_add.append(models.Curriculum(
                    block_id=curriculum_block.id,
                    code=code, name=name, units=units, type=ctype,
                    department_id=target_dept_id, program_code=program_code,
                    year_level=current_year_context, semester_term=current_sem_context,
                    lec_units=lec_units, lab_units=lab_units, pre_requisite=pre_req,
                    is_major=is_major_val
                ))
            all_parsed_data.append(item_data)
            print(f"DEBUG: Successfully captured subject: {code}")

        # STEP 6 — SUMMARY OF UNITS (VALIDATION ONLY)
        total_units_extracted = sum(item["units"] for item in all_parsed_data)
        excel_grand_total = 0
        summary_labels_values = []

        found_summary_start = -1
        for i, row in enumerate(raw_rows):
            row_text = " ".join([str(v).strip().upper() for v in row if not is_na(v)])
            if "SUMMARY OF UNITS" in row_text or "TOTAL UNITS" in row_text:
                found_summary_start = i
                break
        
        if found_summary_start != -1:
            for i in range(found_summary_start, len(raw_rows)):
                s_row = raw_rows[i]
                s_row_text = " ".join([str(v).strip().upper() for v in s_row if not is_na(v)])
                numeric_vals = [v for v in s_row if isinstance(v, (int, float)) and v > 0]
                label_vals = [v.strip() for v in s_row if isinstance(v, str) and len(v.strip()) > 2]
                if numeric_vals:
                    val = numeric_vals[-1]
                    label = label_vals[0] if label_vals else "Unnamed Category"
                    summary_labels_values.append({"label": label, "value": val})
                    if "TOTAL" in s_row_text or i == len(raw_rows) - 1:
                        excel_grand_total = val

        # STEP 8 — OUTPUT STRUCTURE (FULLY DYNAMIC)
        # Group subjects into structured zones as required by Rule 282
        structured_zones = []
        zone_map = {}
        for item in all_parsed_data:
            zone_key = (item["year_level"], item["semester_term"])
            if zone_key not in zone_map:
                zone_map[zone_key] = {
                    "year": item["year_level"],
                    "semester": item["semester_term"],
                    "subjects": [],
                    "zone_total": 0
                }
                structured_zones.append(zone_map[zone_key])
            
            zone_map[zone_key]["subjects"].append(item)
            zone_map[zone_key]["zone_total"] += item["units"]

        validation_status = "Match"
        if excel_grand_total > 0 and abs(total_units_extracted - excel_grand_total) > 0.5:
            validation_status = f"Discrepancy: Excel says {excel_grand_total}, Parser found {total_units_extracted}"

        summary = {
            "program_name": extracted_program_name,
            "academic_year": extracted_ay,
            "selected_sheet": best_sheet,
            "available_sheets": wb.sheetnames,
            "sheet_auto_selected": sheet_auto_selected,
            "selected_sheet_reason": (
                f"Reading '{best_sheet}' because it has Course Code, Title, Units and Lec/Lab columns (confidence {max_score})."
                if sheet_auto_selected
                else f"Reading '{best_sheet}' because you chose it."
            ),
            "scanned_rows": len(raw_rows),
            "total_rows": len(all_parsed_data),
            "detected_subjects": len(all_parsed_data),
            "created_subjects": len(items_to_add),
            "valid_new_items": len(items_to_add),
            "duplicates_skipped": len(skipped_items),
            "skipped_rows": len(skipped_items),
            "issues_found": sum(1 for i in all_parsed_data if i["validation_issues"]),
            "warnings_count": sum(1 for i in all_parsed_data if i["validation_issues"]),
            "errors_count": len(skipped_items),
            "unit_validation": validation_status,
            "excel_total": excel_grand_total,
            "parser_total": total_units_extracted,
            "summary_details": summary_labels_values
        }

        if dry_run:
            return {
                "is_dry_run": True,
                "message": f"Dry run complete. {summary['issues_found']} items have potential issues. {validation_status}",
                "summary": summary,
                "zones": structured_zones, # Structured output (Step 8)
                "report": all_parsed_data, # Flat list for Dayao's Frontend compatibility
                "errors": skipped_items,
                "course": program_code
            }

        if items_to_add:
            db.add_all(items_to_add)
            db.commit()
            
            log_activity(db, current_user.id, "Import Curriculum", f"Imported {len(items_to_add)} items for {extracted_program_name}", "success", department_id=target_dept_id) # type: ignore
            
            return {
                "is_dry_run": False, 
                "message": f"Successfully imported {len(items_to_add)} items", 
                "summary": summary, 
                "zones": structured_zones,
                "errors": skipped_items,
                "course": program_code
            }
        else:
            return {
                "is_dry_run": False, 
                "message": "No new items to import", 
                "summary": summary, 
                "zones": structured_zones,
                "errors": skipped_items,
                "course": program_code
            }

    except HTTPException:
        # A deliberate 4xx raised inside this block -- an unknown sheet name,
        # a missing department -- is a message for the user, not a crash. It was
        # being caught below and re-raised as a 500 with the status code
        # stringified into the detail ("Import failed: 400: ...").
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

@router.post("/import", response_model=schemas.ImportResponse)
async def import_curriculum(
    file: UploadFile = File(...),
    department_id: Optional[int] = Form(None),
    program_code: Optional[str] = Form(None),
    dry_run: bool = Form(False),
    mapping: Optional[str] = Form(None),
    selected_sheet: Optional[str] = Form(None),
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    contents = await file.read()
    return await _process_curriculum_import(
        contents, department_id, program_code, dry_run, db, current_user, mapping, selected_sheet
    )

@router.post("/bulk", response_model=schemas.ImportResponse)
def bulk_create_curriculum(
    payload: schemas.BulkImportRequest,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only System Administrators can bulk import curriculum data")

    try:
        # Step 0: Ensure Curriculum Block exists or create it
        block = None
        if payload.block_id:
            block = db.query(models.CurriculumBlock).filter(models.CurriculumBlock.id == payload.block_id).first()

        if not block:
            block = db.query(models.CurriculumBlock).filter(
                models.CurriculumBlock.program_name == payload.program_name,
                models.CurriculumBlock.academic_year == payload.academic_year,
                models.CurriculumBlock.department_id == payload.department_id
            ).first()

        if not block:
            block = models.CurriculumBlock(
                program_name=payload.program_name,
                academic_year=payload.academic_year,
                department_id=payload.department_id,
                filename="Bulk Import"
            )
            db.add(block)
            db.flush()

        new_items = []
        skipped_items = []
        for item in payload.items:
            # Check if code already exists IN THIS BLOCK
            existing = db.query(models.Curriculum).filter(
                models.Curriculum.block_id == block.id,
                func.lower(models.Curriculum.code) == item.code.lower()
            ).first()

            if not existing:
                new_item = models.Curriculum(**item.model_dump())
                new_item.block_id = block.id
                new_items.append(new_item)
            else:
                skipped_items.append({"code": item.code, "name": item.name, "reason": "Already exists in this block"})
        
        if new_items:
            db.add_all(new_items)
            db.commit()
            
            log_activity(db, current_user.id, "Bulk Create Curriculum", f"Bulk created {len(new_items)} items for {payload.program_name}", "success", department_id=payload.department_id)
        
        summary = {
            "program_name": payload.program_name,
            "academic_year": payload.academic_year,
            "total_rows": len(payload.items),
            "valid_new_items": len(new_items),
            "duplicates_skipped": len(skipped_items),
            "issues_found": 0
        }
        return {
            "is_dry_run": False,
            "message": f"Successfully committed {len(new_items)} items to {payload.program_name}.",
            "summary": summary,
            "report": None,
            "errors": skipped_items
        }
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Bulk commit failed: {str(e)}")

@router.post("/headers")
async def get_excel_headers(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only System Administrators can preview Excel headers")
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_rows = [list(r) for r in list(ws.iter_rows(values_only=True))[:10]]
            if not raw_rows: continue
            rows = []
            for row in raw_rows:
                rows.append([str(v) if v is not None and not (isinstance(v, float) and math.isnan(v)) else "" for v in row])
            sheets_data[sheet_name] = {"sample_rows": rows, "sheet_name": sheet_name}
        return sheets_data
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

@router.delete("/course/{course_name}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_curriculum_course(
    course_name: str,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only System Administrators can delete curriculum courses")
    db.query(models.Curriculum).filter(models.Curriculum.program_code == course_name).delete(synchronize_session=False)
    db.commit()
    
    # Get department for logging
    dept = db.query(models.Department).filter(
        (models.Department.code == current_user.department) | 
        (models.Department.name == current_user.department)
    ).first()
    log_activity(db, current_user.id, "Delete Course", f"Deleted all items for course: {course_name}", "success", department_id=dept.id if dept else None) # type: ignore
    
    return Response(status_code=status.HTTP_204_NO_CONTENT)

@router.delete("/block/{block_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_curriculum_block(
    block_id: int,
    db: Session = Depends(database.get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role != 'admin':
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only System Administrators can delete curriculum blocks")
    
    # Get block details for logging before deletion
    block = db.query(models.CurriculumBlock).filter(models.CurriculumBlock.id == block_id).first()
    block_name = block.program_name if block else str(block_id)
    dept_id = block.department_id if block else None

    # Delete all subjects in the block first
    db.query(models.Curriculum).filter(models.Curriculum.block_id == block_id).delete(synchronize_session=False)
    # Then delete the block itself
    db.query(models.CurriculumBlock).filter(models.CurriculumBlock.id == block_id).delete(synchronize_session=False)
    
    db.commit()
    
    log_activity(db, current_user.id, "Delete Block", f"Deleted curriculum block: {block_name}", "success", department_id=dept_id) # type: ignore
    
    return Response(status_code=status.HTTP_204_NO_CONTENT)

"""
A second DELETE /{id} handler sat here, duplicating the one above.
FastAPI matches the first registration, so it never ran — and it skipped the
audit-log write the live handler performs.
"""
